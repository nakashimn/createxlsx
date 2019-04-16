import itertools
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.chart import ScatterChart
from openpyxl.chart import BarChart
from openpyxl.chart import Reference
from openpyxl.chart import Series


def copy_dataframe_to_worksheet(wb, dataframe, title=None):
    """
    @fn copy_dataframe_to_worksheet()
    @brief
    @param wb コピー先ワークブック
    @param dataframe コピー元データフレーム(pandas.DataFrame)
    @param title シート名 =<Active Sheet>
    @retval ws ワークシート
    """
    if title is None:
        ws = wb.active
    else:
        ws = wb.create_sheet(title=title)
    for r in dataframe_to_rows(dataframe, index=False, header=True):
        ws.append(r)
    return ws


def copy_dataframe_to_cells(ws, dataframe, start_cell="A1"):
    """
    @fn copy_dataframe_to_cells()
    @brief
    @param ws コピー先シート
    @param dataframe コピー元データフレーム(pandas.DataFrame)
    @param start_cell コピー先セル(左上) ="A1"
    @retval cells コピー先セル範囲
    """
    start_row = ws[start_cell].row
    start_col = ws[start_cell].col_idx

    for row, index in enumerate(dataframe.index):
        ws.cell(row=row+start_row+1, column=start_col).value = index

    for col, header in enumerate(dataframe.columns):
        ws.cell(row=start_row, column=col+start_col+1).value = header

    for (row, index), (col, header) in itertools.product(enumerate(dataframe.index), enumerate(dataframe.columns)):
        ws.cell(row=row+start_row+1, column=col+start_col+1).value = dataframe.iloc[row, col]

    cells = Reference(ws,
                      min_row=start_row+1,
                      max_row=start_row+len(dataframe.index),
                      min_col=start_col+1,
                      max_col=start_row+len(dataframe.columns))
    return cells


def create_scatter_chart(x_cells, y_cells, x_title, y_title, x_range=None, y_range=None, legends=None, height=10, width=20):
    """
    @fn create_scatter_chart()
    @brief
    @param x_cells 横軸データ参照範囲(Reference)
    @param y_cells 縦軸データ参照範囲(Reference)
    @param x_title 横軸ラベル
    @param y_title 縦軸ラベル
    @param x_range 定義域
    @param y_range 値域
    @param legends 凡例
    @param height グラフの高さ
    @param width グラフの幅
    @retval chart グラフ
    """
    chart = ScatterChart()
    chart.x_axis.title = x_title
    chart.y_axis.title = y_title
    chart.style = 2
    chart.height = height
    chart.width = width
    if x_range is not None:
        chart.x_axis.scaling.min = min(x_range)
        chart.x_axis.scaling.max = max(x_range)
    if y_range is not None:
        chart.y_axis.scaling.min = min(y_range)
        chart.y_axis.scaling.max = max(y_range)
    if legends is None:
        chart.legend = None
    else:
        chart.legend.position = "t"
    if type(x_cells) != list and type(y_cells) != list:
        series = Series(y_cells, x_cells, title=legends)
        chart.series.append(series)
    elif type(x_cells) != list and type(y_cells) == list:
        for y_cells_unit, legend in zip(y_cells, legends):
            series = Series(y_cells_unit, x_cells, title=legend)
            chart.series.append(series)
    elif type(x_cells) == list and type(y_cells) == list:
        for x_cells_unit, y_cells_unit, legend in zip(x_cells, y_cells, legends):
            series = Series(y_cells_unit, x_cells_unit, title=legend)
            chart.series.append(series)
    return chart


def create_bar_chart(x_cells, y_cells, title, range=None, height=10, width=20):
    """
    @fn create_bar_chart()
    @brief
    @param x_cells カテゴリ参照範囲(Reference)
    @param y_cells データ参照範囲(Reference)
    @param title タイトル
    @param range 値域
    @param height グラフの高さ
    @param width グラフの幅
    @retval chart グラフ
    """
    chart = BarChart()
    chart.title = title
    chart.style = 2
    chart.height = height
    chart.width = width
    chart.legend = None
    chart.add_data(y_cells)
    chart.set_categories(x_cells)
    return chart
